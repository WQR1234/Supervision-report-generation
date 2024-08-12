from docxtpl import DocxTemplate
from docxtpl import InlineImage
import openpyxl
import pprint
import os
from PIL import Image
from datetime import datetime, timedelta
from docx.shared import Mm
import re
import urllib.request


pic_id = 0


def is_url(s):
    # 正则表达式来匹配URL
    regex = r'^(?:http|ftp)s?://' # 支持http和https，以及ftp和ftps
    regex += r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?)|' # 域名
    regex += r'localhost|' #  localhost
    regex += r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})' # IP地址
    regex += r'(?::\d+)?' # 可选的端口
    regex += r'(?:/?|[/?]\S+)$' # 路径
    regex = re.compile(regex)

    return re.match(regex, s) is not None


def compress_picture(image_path, max_size=1024 * 1024):
    # 调整图片尺寸为8cm x 6cm
    image = Image.open(image_path)
    width, height = image.size
    aspect_ratio = width / height
    new_width = 8 * 72  # 8cm转换为磅
    new_height = new_width / aspect_ratio
    if new_height > 6 * 72:  # 6cm转换为磅
        new_height = 6 * 72
        new_width = new_height * aspect_ratio
    image = image.resize((int(new_width), int(new_height)))

    # 压缩图片
    quality = 85
    while os.path.getsize(image_path) > max_size and quality >= 5:
        image.save(image_path, optimize=True, quality=quality)

    return image_path


def trans_date(dictionary):
    for key, value in dictionary.items():
        if isinstance(value, dict):
            trans_date(value)  # 递归调用process_dict函数来处理子字典
        elif isinstance(value, list):
            for item in value:
                if isinstance(item, dict):
                    trans_date(item)  # 递归调用process_dict函数来处理字典
        elif isinstance(key, str) and '日期' in key:
            try:
                date_value = datetime(1900, 1, 1) + timedelta(days=value)
                formatted_date = date_value.strftime('%Y年%m月%d日')  # 将日期格式化为yyyy年mm月dd日的字符串
                dictionary[key] = formatted_date  # 更新键值为格式化后的日期字符串
            except ValueError:
                pass  # 忽略无效日期数字


def generate_dic(xls_path):
    wb = openpyxl.load_workbook(xls_path)
    sheet_names = wb.sheetnames
    result_dict = {}

    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        rows = sheet.iter_rows(min_row=1, values_only=True)
        headers = next(rows)  # 第一行作为字典的键

        data_list = []
        for row in rows:
            row_dict = {}
            for i, value in enumerate(row):
                row_dict[headers[i]] = value
            data_list.append(row_dict)

        result_dict[sheet_name] = data_list

    # pprint.pprint(result_dict)
    return result_dict


def formate_dic(dic, pic_path, tpl):
    empty_keys = []
    for key, value in dic.items():
        if isinstance(value, dict):
            formate_dic(value)  # 递归调用process_dict函数来处理子字典
        elif isinstance(value, list):
            for item in value:
                if isinstance(item, dict):
                    formate_dic(item, pic_path, tpl)  # 递归调用process_dict函数来处理字典
        else:
            if value is None:
                empty_keys.append(key)
                continue
            # 格式化日期
            if isinstance(key, str) and '日期' in key:
                try:
                    date_value = datetime(1900, 1, 1) + timedelta(days=value)
                    formatted_date = date_value.strftime('%Y年%m月%d日')  # 将日期格式化为yyyy年mm月dd日的字符串
                    dic[key] = formatted_date  # 更新键值为格式化后的日期字符串
                except ValueError:
                    pass  # 忽略无效日期数字
            # 格式化图片
            elif isinstance(key, str) and '图片' in key:
                is_pic = False
                if is_url(value):
                    is_pic = True
                    global pic_id
                    pic_name = os.path.join(pic_path, str(pic_id) + '.jpg')
                    pic_id += 1
                    urllib.request.urlretrieve(value, pic_name)
                if value.lower().endswith(('.jpg', '.jpeg', '.png')):
                    is_pic = True
                    pic_name = os.path.join(pic_path, value)
                if is_pic:
                    compress_picture(pic_name)
                    dic[key] = InlineImage(tpl, image_descriptor=pic_name, width=Mm(60))
            # 格式化一般段落
            else:
                if isinstance(value, str):
                    dic[key] = value.replace('\n', '\a')
    
    for k in empty_keys:
        del dic[k]


def generate_file(mode, new_name, data, pic_path='Temp'):
    tpl = DocxTemplate(mode)
    if isinstance(data, str):
        dic = generate_dic(data)
    if isinstance(data, dict):
        dic = data
    formate_dic(dic, pic_path, tpl)
    # print()
    # pprint.pprint(dic)
    tpl.render(dic)
    tpl.save(new_name)


if __name__ == "__main__":
    generate_file('The_group/交流监造联系单.docx', 'The_group/交流监造联系单-0.docx', 'The_group/副本交流监造联系单.xlsx')
